import openpyxl
from openpyxl.styles import Font, Alignment

from utils.tools import get_now_time_num, get_now_time_format


class CameraExcel:
    def __init__(self, case_id_column, result_column, sheet_index: int = 0,
                 template_path='../template/PAN_v2_测试用例.xlsx'):
        """
        case_id_column: case id的列 e.g 'A'
        result_column: 写入结果的列  e.g 'D'
        sheet_index: sheet的索引 e.g 0
        template_path: 模板用例的地址
        """
        self.case_id_column = case_id_column
        self.result_column = result_column
        self.sheet_index = sheet_index
        wb = openpyxl.load_workbook(template_path)
        self.title_font = Font(u'宋体', size=16, bold=True, color='000000')
        self.info_font = Font(u'微软雅黑', size=12, bold=True, color='000000')
        self.header_font = Font(u'微软雅黑', size=10, bold=True, color='8B0000')
        self.header_align = Alignment(horizontal='center', vertical='center')
        self.pass_font = Font(u'微软雅黑', size=9, bold=True, color='000000')
        self.fail_font = Font(u'微软雅黑', size=9, bold=True, color='FF0000')
        self.file_path = '../excel/' + 'PAN_v2_测试用例' + \
                         get_now_time_num() + '.xlsx'
        sheet_names = wb.sheetnames
        sheet = wb[sheet_names[self.sheet_index]]
        sheet['C9'] = get_now_time_format()
        sheet['E9'] = 'ui_auto_tester'
        wb.save(self.file_path)

    def inset_case_result(self, case_id, result, check=False):
        """写入case的结果"""
        row = self.get_cell_row(self.case_id_column, case_id)
        if row is None:
            raise ValueError("未搜索到该case id %s" % case_id)
        location = self.result_column + str(row)
        print("location:", location)
        wb = openpyxl.load_workbook(self.file_path)
        sheet_names = wb.sheetnames
        sheet = wb[sheet_names[self.sheet_index]]
        print(sheet)
        if result in ['pass', 'P', 'Pass', 'PASS', True]:
            if check:
                sheet[location] = 'Check'
            else:
                sheet[location] = 'Pass'
            sheet[location].font = self.pass_font
        elif result in ['fail', 'F', 'Fail', 'FAIL', False]:
            if check:
                sheet[location] = 'Fail(check)'
            else:
                sheet[location] = 'Fail'
            sheet[location].font = self.fail_font
        else:
            raise ValueError("case result error")
        wb.save(self.file_path)

    def get_cell_row(self, cell_column, cell_value):
        """根据内容获取行数"""
        wb = openpyxl.load_workbook(self.file_path)
        sheet_names = wb.sheetnames
        sheet = wb[sheet_names[self.sheet_index]]
        i = [i for i in sheet[cell_column] if i.value == cell_value]
        wb.close()
        if len(i) > 1:
            raise ValueError("在excel中检索到多个" + str(cell_value), str([i]))
        return i[0].row if len(i) == 1 else None
